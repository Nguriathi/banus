import streamlit as st
import pandas as pd
import plotly.express as px
from io import BytesIO
import json
from streamlit_lottie import st_lottie
from prophet import Prophet
import openpyxl
import os

# PAGE CONFIG (must be first!)
st.set_page_config(page_title="Product Invoice Analyzer", layout="wide", page_icon="📦")

# SPLASH SCREEN
if "splash_shown" not in st.session_state:
    st.session_state.splash_shown = False

if not st.session_state.splash_shown:
    st.markdown("""
    <style>
    .splash {
        text-align: center;
        padding: 4rem 1rem;
        background: linear-gradient(135deg, #4f8bf9, #3a6db0);
        color: white;
        border-radius: 20px;
        margin: 2rem 0;
    }
    </style>
    """, unsafe_allow_html=True)
    st.markdown("""
    <div class="splash">
        <h1>📦 Enterprise Product Invoice Analyzer</h1>
        <p>Visualize, summarize, and explore your invoice data with advanced analytics.</p>
    </div>
    """, unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        if st.button("🚀 Start Now", use_container_width=True):
            st.session_state.splash_shown = True
    st.stop()

# SIDEBAR LOTTIE & MODE SELECTOR
def load_lottiefile(filepath: str):
    with open(filepath, "r") as f:
        return json.load(f)

with st.sidebar:
    st.markdown("""
    <div style="text-align:center;font-size:2.4rem;font-weight:bold;letter-spacing:0.12em;margin-bottom:0.5rem;">
        IA
    </div>
    """, unsafe_allow_html=True)
    lottie_json = load_lottiefile("assets/animation.json")
    st_lottie(lottie_json, height=170, key="lottie-animation")
    st.markdown("### 📊 Product Invoice Analyzer")
    st.write("Upload your invoice Excel file(s) and instantly analyze your product data.")
    mode = st.radio("Select Processing Mode", ["Single File Processing", "Batch Processing (Multiple Files)"])

# MAIN PAGE
st.markdown("""
    <h1 style="text-align:center;font-size:2.3rem;font-weight:700;margin-bottom:1.5rem;">
        Invoice Analyzer
    </h1>
""", unsafe_allow_html=True)
st.markdown("## 🗂️ Invoice File Upload")

if mode == "Single File Processing":
    uploaded_files = st.file_uploader(
        "Upload a single Excel file with an 'analysis' sheet",
        type=["xlsx"],
        accept_multiple_files=False,
        key="single_file"
    )
else:
    uploaded_files = st.file_uploader(
        "Upload multiple Excel files (all must have an 'analysis' sheet)",
        type=["xlsx"],
        accept_multiple_files=True,
        key="batch_files"
    )

def extract_metadata(df):
    agent, vessel = None, None
    for i in range(len(df)):
        row = df.iloc[i].astype(str)
        if row[0].strip().upper() == "AGENT":
            agent = row[1].strip()
        if row[0].strip().upper() == "VESSEL":
            vessel = row[1].strip()
    return agent, vessel

def extract_product_table(df):
    start_idx = df[df.iloc[:, 0].astype(str).str.strip().str.upper() == 'NO'].index
    if len(start_idx) == 0:
        return pd.DataFrame()
    start = start_idx[0] + 1
    end = start
    while end < len(df) and pd.to_numeric(df.iloc[end, 0], errors='coerce') == df.iloc[end, 0]:
        end += 1
    table = df.iloc[start:end, :6].copy()
    table.columns = ['NO', 'PRODUCT DESCRIPTION', 'UNIT/PRC', 'UNIT', 'QTY', 'TOTAL USD']
    table = table.dropna(subset=['PRODUCT DESCRIPTION'])
    table['QTY'] = pd.to_numeric(table['QTY'], errors='coerce')
    table['TOTAL USD'] = pd.to_numeric(table['TOTAL USD'], errors='coerce')
    table['UNIT/PRC'] = pd.to_numeric(table['UNIT/PRC'], errors='coerce')
    table = table.dropna(subset=['QTY', 'TOTAL USD', 'UNIT/PRC'])
    return table

def extract_invoice_date(df):
    dod_row = df[df.iloc[:, 0].astype(str).str.strip().str.upper() == "DOD"]
    if not dod_row.empty:
        return pd.to_datetime(dod_row.iloc[0, 1], errors='coerce')
    return pd.NaT

def prophet_forecast(ts_df, periods=3):
    ts = ts_df.groupby('INVOICE_DATE')['QTY'].sum().reset_index()
    ts = ts.rename(columns={'INVOICE_DATE': 'ds', 'QTY': 'y'})
    if len(ts) < 2:
        return None
    model = Prophet(yearly_seasonality=False, weekly_seasonality=False, daily_seasonality=False)
    model.fit(ts)
    future = model.make_future_dataframe(periods=periods, freq='MS')
    forecast = model.predict(future)
    return forecast

def fill_template(template_path, agent, vessel, table_df):
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    ws['A2'] = agent
    ws['B2'] = vessel
    start_row = 10
    for i, row in table_df.iterrows():
        ws.cell(row=start_row+i, column=1, value=row['NO'])
        ws.cell(row=start_row+i, column=2, value=row['PRODUCT DESCRIPTION'])
        ws.cell(row=start_row+i, column=3, value=row['UNIT/PRC'])
        ws.cell(row=start_row+i, column=4, value=row['UNIT'])
        ws.cell(row=start_row+i, column=5, value=row['QTY'])
        ws.cell(row=start_row+i, column=6, value=row['TOTAL USD'])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# NAVBAR CSS (gradient background)
st.markdown("""
<style>
.navbar-summary {
    background: linear-gradient(135deg, #4f8bf9, #3a6db0);
    color: white !important;
    border-radius: 16px;
    padding: 1.2rem 2rem;
    display: flex;
    justify-content: space-around;
    margin-bottom: 1.5rem;
    font-size: 1.25rem;
}
.navbar-summary > div {
    text-align: center;
}
.navbar-summary .big {
    font-size: 2.1rem;
    font-weight: bold;
    color: #fff;
}
</style>
""", unsafe_allow_html=True)

# --- SINGLE FILE PROCESSING ---
if mode == "Single File Processing" and uploaded_files:
    df = pd.read_excel(uploaded_files, sheet_name="analysis", header=None)
    agent, vessel = extract_metadata(df)
    product_table = extract_product_table(df)
    if not product_table.empty:
        st.markdown(f"""
        <div style="text-align:center;font-size:1.3rem;font-weight:600;">
            Vessel: <span style="color:#4f8bf9">{vessel or "N/A"}</span> | 
            Agent: <span style="color:#4f8bf9">{agent or "N/A"}</span>
        </div>
        """, unsafe_allow_html=True)
        total_qty = int(product_table['QTY'].sum())
        total_value = float(product_table['TOTAL USD'].sum())
        num_items = product_table['PRODUCT DESCRIPTION'].nunique()
        st.markdown(
            f"""
            <div class="navbar-summary">
                <div>
                    <div class="big">{total_qty:,}</div>
                    <div>Total Quantity</div>
                </div>
                <div>
                    <div class="big">${total_value:,.2f}</div>
                    <div>Total Value</div>
                </div>
                <div>
                    <div class="big">{num_items}</div>
                    <div>Unique Items</div>
                </div>
            </div>
            """, unsafe_allow_html=True
        )

        st.markdown("### ✏️ Edit Product Table Below")
        edited_table = st.data_editor(
            product_table,
            column_config={
                "UNIT/PRC": st.column_config.NumberColumn(format="$%.2f"),
                "TOTAL USD": st.column_config.NumberColumn(format="$%.2f"),
                "QTY": st.column_config.NumberColumn(format="%d"),
            },
            num_rows="dynamic"
        )

        # --- Chart analysis and summary ---
        st.markdown("### 📊 Product Description Visualization")
        chart_data = edited_table.groupby("PRODUCT DESCRIPTION").agg(
            {"QTY": "sum", "TOTAL USD": "sum"}
        ).reset_index()
        n_items = len(chart_data)
        chart_summary = []
        if n_items <= 10:
            fig = px.bar(
                chart_data,
                x="PRODUCT DESCRIPTION",
                y="QTY",
                color="QTY",
                text="QTY",
                title="Quantity by Product",
            )
            chart_summary.append(f"- Top product: **{chart_data.sort_values('QTY', ascending=False).iloc[0]['PRODUCT DESCRIPTION']}**")
            fig.update_traces(textposition='outside')
            fig.update_layout(xaxis_tickangle=-45, height=500)
            st.plotly_chart(fig, use_container_width=True)
        elif n_items <= 25:
            fig = px.bar(
                chart_data.sort_values("QTY"),
                x="QTY",
                y="PRODUCT DESCRIPTION",
                orientation="h",
                color="QTY",
                text="QTY",
                title="Quantity by Product",
            )
            chart_summary.append(f"- Top product: **{chart_data.sort_values('QTY', ascending=False).iloc[0]['PRODUCT DESCRIPTION']}**")
            fig.update_layout(height=600)
            st.plotly_chart(fig, use_container_width=True)
        else:
            fig = px.treemap(
                chart_data,
                path=["PRODUCT DESCRIPTION"],
                values="QTY",
                color="TOTAL USD",
                color_continuous_scale="Blues",
                title="Product Quantity Treemap",
            )
            chart_summary.append(f"- Most significant product by quantity: **{chart_data.sort_values('QTY', ascending=False).iloc[0]['PRODUCT DESCRIPTION']}**")
            st.plotly_chart(fig, use_container_width=True)

        # --- Download as filled template ---
        if os.path.exists("template.xlsx"):
            templ_out = fill_template("template.xlsx", agent, vessel, edited_table)
            st.download_button(
                label="📥 Download Filled Invoice Template",
                data=templ_out,
                file_name=f"invoice_{vessel or 'vessel'}_{agent or 'agent'}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("Template file 'template.xlsx' not found in the current directory.")

        st.markdown("### 📝 Summary")
        summary_bullets = [
            f"- This invoice contains **{num_items} unique products**.",
            f"- Total quantity: **{total_qty}**.",
            f"- Total value: **${total_value:,.2f}**.",
        ] + chart_summary
        st.markdown("\n".join(summary_bullets))
    else:
        st.warning("No valid product table found in your file.")

# --- BATCH PROCESSING ---
elif mode == "Batch Processing (Multiple Files)" and uploaded_files:
    all_tables = []
    for file in uploaded_files:
        df = pd.read_excel(file, sheet_name="analysis", header=None)
        product_table = extract_product_table(df)
        invoice_date = extract_invoice_date(df)
        if not product_table.empty:
            product_table["INVOICE_DATE"] = invoice_date
            product_table["SOURCE_FILE"] = file.name
            all_tables.append(product_table)
    if all_tables:
        combined = pd.concat(all_tables, ignore_index=True)
        st.markdown("### 🗃️ Combined Product Table (Editable)")
        edited_combined = st.data_editor(
            combined,
            column_config={
                "UNIT/PRC": st.column_config.NumberColumn(format="$%.2f"),
                "TOTAL USD": st.column_config.NumberColumn(format="$%.2f"),
                "QTY": st.column_config.NumberColumn(format="%d"),
                "INVOICE_DATE": st.column_config.DateColumn(),
            },
            num_rows="dynamic"
        )

        total_qty = int(edited_combined['QTY'].sum())
        total_value = float(edited_combined['TOTAL USD'].sum())
        num_items = edited_combined['PRODUCT DESCRIPTION'].nunique()
        st.markdown(
            f"""
            <div class="navbar-summary">
                <div>
                    <div class="big">{total_qty:,}</div>
                    <div>Total Quantity</div>
                </div>
                <div>
                    <div class="big">${total_value:,.2f}</div>
                    <div>Total Value</div>
                </div>
                <div>
                    <div class="big">{num_items}</div>
                    <div>Unique Items</div>
                </div>
            </div>
            """, unsafe_allow_html=True
        )

        # --- Analysis charts and summaries ---
        summaries = [
            f"- Processed **{len(uploaded_files)} files**.",
            f"- Combined total of **{num_items} unique products**.",
            f"- Total quantity: **{total_qty}**.",
            f"- Total value: **${total_value:,.2f}**."
        ]

        # Top products (bar)
        st.markdown("### 🥇 Top Products (Bar Chart)")
        most_bought = edited_combined.groupby("PRODUCT DESCRIPTION")["QTY"].sum().sort_values(ascending=False).reset_index()
        fig_most = px.bar(
            most_bought.head(10),
            x="PRODUCT DESCRIPTION",
            y="QTY",
            color="QTY",
            text="QTY",
            title="Top 10 Most Bought Products"
        )
        top_product = most_bought.iloc[0]['PRODUCT DESCRIPTION']
        summaries.append(f"- Top product across all files: **{top_product}** with quantity **{most_bought.iloc[0]['QTY']}**.")
        fig_most.update_traces(textposition='outside')
        fig_most.update_layout(xaxis_tickangle=-45)
        st.plotly_chart(fig_most, use_container_width=True)

        # Treemap
        st.markdown("### 🟦 Product Composition (Treemap)")
        fig_treemap = px.treemap(
            most_bought.head(20),
            path=["PRODUCT DESCRIPTION"],
            values="QTY",
            color="QTY",
            color_continuous_scale="Blues",
            title="Product Share in Total Quantity"
        )
        summaries.append(f"- Product with largest share: **{most_bought.iloc[0]['PRODUCT DESCRIPTION']}**.")
        st.plotly_chart(fig_treemap, use_container_width=True)

        # Quantity trends
        st.markdown("### 📈 Quantity Trends (Line Chart)")
        qty_trend = edited_combined.groupby(["INVOICE_DATE", "PRODUCT DESCRIPTION"])["QTY"].sum().reset_index()
        fig_line = px.line(
            qty_trend,
            x="INVOICE_DATE",
            y="QTY",
            color="PRODUCT DESCRIPTION",
            title="Product Quantity Trends Over Time"
        )
        st.plotly_chart(fig_line, use_container_width=True)
        summaries.append("- Displayed quantity trends for all products over time.")

        # Price trends
        st.markdown("### 💸 Price Trends (Line Chart)")
        price_trend = edited_combined.groupby(["INVOICE_DATE", "PRODUCT DESCRIPTION"])["UNIT/PRC"].mean().reset_index()
        fig_price = px.line(
            price_trend,
            x="INVOICE_DATE",
            y="UNIT/PRC",
            color="PRODUCT DESCRIPTION",
            title="Unit Price Trends by Product"
        )
        st.plotly_chart(fig_price, use_container_width=True)
        summaries.append("- Displayed price trends for all products over time.")

        # Recurring high-quantity items
        st.markdown("### 🔁 Recurring High-Quantity Items (Bar Chart)")
        recurring = edited_combined.groupby("PRODUCT DESCRIPTION")["QTY"].agg(['count', 'sum']).reset_index()
        recurring = recurring[recurring['count'] > 1].sort_values('sum', ascending=False)
        if not recurring.empty:
            fig_recurring = px.bar(
                recurring.head(10),
                x="PRODUCT DESCRIPTION",
                y="sum",
                color="sum",
                text="sum",
                title="Top Recurring Items (by Total Quantity)"
            )
            top_recurring = recurring.iloc[0]['PRODUCT DESCRIPTION']
            summaries.append(f"- Most recurring high-quantity item: **{top_recurring}**.")
            fig_recurring.update_traces(textposition='outside')
            fig_recurring.update_layout(xaxis_tickangle=-45)
            st.plotly_chart(fig_recurring, use_container_width=True)
        else:
            st.info("No recurring items found in large quantities.")
            summaries.append("- No recurring high-quantity items found.")

        # Individual product trends and forecasting
        st.markdown("### 🔍 Individual Product Trends & Forecast")
        product_options = most_bought["PRODUCT DESCRIPTION"].tolist()
        product_selected = st.selectbox("Select Product for Trend & Prediction", product_options)
        prod_trend = edited_combined[edited_combined["PRODUCT DESCRIPTION"] == product_selected]
        prod_trend_agg = prod_trend.groupby("INVOICE_DATE")["QTY"].sum().reset_index()
        fig_prod_trend = px.line(
            prod_trend_agg,
            x="INVOICE_DATE",
            y="QTY",
            markers=True,
            title=f"Quantity Trend for {product_selected}"
        )
        st.plotly_chart(fig_prod_trend, use_container_width=True)
        summaries.append(f"- Displayed quantity trend for **{product_selected}**.")

        # Prophet forecast
        if prod_trend["INVOICE_DATE"].notnull().sum() >= 2:
            forecast = prophet_forecast(prod_trend[["INVOICE_DATE", "QTY"]], periods=3)
            if forecast is not None:
                fig_pred = px.line(
                    forecast,
                    x="ds",
                    y="yhat",
                    title=f"Forecasted Quantity for {product_selected} (Next 3 Periods)"
                )
                fig_pred.add_scatter(x=forecast["ds"], y=forecast["yhat_lower"], mode='lines', name='Lower CI', line=dict(dash='dash'))
                fig_pred.add_scatter(x=forecast["ds"], y=forecast["yhat_upper"], mode='lines', name='Upper CI', line=dict(dash='dash'))
                st.plotly_chart(fig_pred, use_container_width=True)
                summaries.append(f"- Forecasted future quantities for **{product_selected}**.")
            else:
                st.info("Not enough data for forecasting this product.")
                summaries.append(f"- Not enough data for forecasting **{product_selected}**.")
        else:
            st.info("Not enough data for forecasting this product.")
            summaries.append(f"- Not enough data for forecasting **{product_selected}**.")

        # Scatter plot: Outlier detection
        st.markdown("### 🟠 Outlier Detection (Scatter Plot: Price vs Quantity)")
        fig_scatter = px.scatter(
            prod_trend,
            x="QTY",
            y="UNIT/PRC",
            size="TOTAL USD",
            color="INVOICE_DATE",
            title=f"Scatter: Price vs Quantity for {product_selected}",
            hover_data=["INVOICE_DATE"]
        )
        st.plotly_chart(fig_scatter, use_container_width=True)
        summaries.append(f"- Outlier detection scatter for **{product_selected}**.")

        # Download
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            edited_combined.to_excel(writer, index=False, sheet_name='Batch_Product_Table')
        output.seek(0)
        st.download_button(
            label="📥 Download Combined Table as Excel",
            data=output,
            file_name="batch_invoice_table.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        st.markdown("### 📝 Batch Summary")
        st.markdown("\n".join(summaries))
    else:
        st.warning("No valid product tables found in the uploaded files.")

elif mode == "Single File Processing" or mode == "Batch Processing (Multiple Files)":
    st.info("Upload your Excel file(s) to begin.")
